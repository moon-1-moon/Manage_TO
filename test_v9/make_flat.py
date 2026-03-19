import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ================= 설정 부분 =================
base_dir    = Path(__file__).parent
input_file  = "25년.xlsx"
output_name = "25년_플랫데이터.xlsx"
# ============================================

HDR_FILL = PatternFill('solid', start_color='1F4E79')
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
KEY_FILL = PatternFill('solid', start_color='D6E4F0')
KEY_FONT = Font(name='Arial', size=10)
NUM_FONT = Font(name='Arial', size=10)
CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT     = Alignment(horizontal='left',   vertical='center')
RIGHT    = Alignment(horizontal='right',  vertical='center')

def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def load_정원(xl, sheet, val_col):
    """정원/현원 시트 로드 — 기관코드 float→int→str 변환 포함"""
    df = pd.read_excel(xl, sheet_name=sheet, header=0)
    df.columns = ['기관코드', '직급_행', val_col]
    df['기관코드'] = df['기관코드'].apply(
        lambda x: '' if (isinstance(x, float) and np.isnan(x))
                  else str(int(x)) if isinstance(x, float)
                  else str(x)
    )
    df = df[df['기관코드'] != '']
    df[val_col] = pd.to_numeric(df[val_col], errors='coerce')
    return df

def get_output_path(base_path, name):
    p = base_path / name
    if not p.exists():
        return p
    stem, ext = p.stem, p.suffix
    i = 2
    while True:
        candidate = base_path / f"{stem}_{i}{ext}"
        if not candidate.exists():
            return candidate
        i += 1

def main():
    xl = pd.ExcelFile(base_dir / input_file)

    # ── 1. LDAP 로드 ──────────────────────────────────────
    df_ldap = pd.read_excel(xl, sheet_name='LDAP(25)', header=0)
    df_ldap.columns = ['기관코드', '기관명2', '기관명3', '기관명4', '기관명5', '기관명6', '기관명_전체']
    df_ldap['기관코드'] = df_ldap['기관코드'].astype(str)

    # ── 2. 정원/현원 로드 ─────────────────────────────────
    df_기준 = load_정원(xl, '(25)기준한시', '기준한시정원')
    df_운영 = load_정원(xl, '(25)운영한시', '운영한시정원')
    df_현원 = load_정원(xl, '(25)현원',    '현원')

    # ── 3. 기관코드+직급_행 기준 집계 후 병합 ─────────────
    KEY = ['기관코드', '직급_행']
    agg = (
        df_기준.groupby(KEY)['기준한시정원'].sum().reset_index()
        .merge(df_운영.groupby(KEY)['운영한시정원'].sum().reset_index(), on=KEY, how='outer')
        .merge(df_현원.groupby(KEY)['현원'].sum().reset_index(),         on=KEY, how='outer')
    )

    # ── 4. LDAP 병합 (right join → 정원 시트 기준 전체 유지) ──
    flat = df_ldap.merge(agg, on='기관코드', how='right')
    flat = flat.sort_values(['기관코드', '직급_행']).reset_index(drop=True)

    # ── 5. 검증 ───────────────────────────────────────────
    print(f"총 행 수              : {len(flat):,}")
    for col, df_src in [('기준한시정원', df_기준), ('운영한시정원', df_운영), ('현원', df_현원)]:
        원본 = df_src[col].sum()
        결과 = flat[col].sum()
        ok   = '✅' if abs(원본 - 결과) < 1e-6 else '❌'
        print(f"{col:12s} 합계  : {결과:.0f}  (원본: {원본:.0f}) {ok}")

    # ── 6. 엑셀 저장 ──────────────────────────────────────
    col_names  = ['기관코드', '기관명2', '기관명3', '기관명4', '기관명5', '기관명6',
                  '기관명_전체', '직급_행', '기준한시정원', '운영한시정원', '현원']
    col_widths = [12, 14, 18, 20, 20, 22, 40, 10, 14, 14, 10]
    # 열 인덱스 기준: 1~7 → 텍스트 키, 8(직급_행) → 텍스트, 9~11 → 숫자
    TEXT_COLS = set(range(1, 9))   # 기관코드~직급_행
    NUM_COLS  = set(range(9, 12))  # 기준한시정원~현원

    output_path = get_output_path(base_dir, output_name)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        flat[col_names].to_excel(writer, index=False, sheet_name='플랫데이터')
        ws = writer.sheets['플랫데이터']

        # 헤더 서식
        for ci, (col, w) in enumerate(zip(col_names, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = CENTER
            cell.border    = thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 22

        # 데이터 서식
        for ri in range(2, len(flat) + 2):
            for ci in range(1, len(col_names) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.border = thin_border()
                if ci in TEXT_COLS:
                    cell.font      = KEY_FONT
                    cell.fill      = KEY_FILL
                    cell.alignment = LEFT
                    if ci == 1:
                        cell.number_format = '@'
                else:
                    cell.font      = NUM_FONT
                    cell.alignment = RIGHT
                    if cell.value is not None:
                        cell.number_format = '#,##0.0'

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(col_names))}1'

    print(f"저장 완료             : {output_path}")

if __name__ == "__main__":
    main()