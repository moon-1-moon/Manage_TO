import openpyxl
from pathlib import Path

FILE_PATH = Path(__file__).parent / "25년 한시정원_데이터.xlsx"

wb = openpyxl.load_workbook(FILE_PATH)
ws = wb.active

header = ws["J1"].value
total_rows = ws.max_row - 1  # 헤더 제외

total = sum(
    ws.cell(row=r, column=13).value
    for r in range(2, ws.max_row + 1)
    if ws.cell(row=r, column=13).value is not None
)

print(f"열 이름   : {header}")
print(f"데이터 행 수: {total_rows:,}행")
print(f"정원 합계  : {int(total):,}")