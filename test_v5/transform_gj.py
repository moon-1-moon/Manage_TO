import openpyxl
import pandas as pd
import os

# File paths
input_file = r'c:\Workspace\Manage_TO\test_v5\(251230)기준정원(본부).xlsx'
output_file = r'c:\Workspace\Manage_TO\test_v5\(251230)기준정원(본부)_데이터.xlsx'

# Load the workbook and select the specific sheet
try:
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb['(251230)기준정원(본부)']
except Exception as e:
    print(f"Error loading workbook or sheet: {e}")
    exit(1)

# Initialize a list to hold the transformed data
data = []

# Define the columns (fixed based on requirements)
columns = [
    '기관코드', '기관명', '기관구분', '소속기관차수', '중분류', 
    '소분류', '기구유형', '보임직급', '상호이체보임직급', 
    '직군', '계급_및_직급', '직렬', '기준정원'
]

# Iterate through the specified range (Rows 11-96, Columns N-BS)
# openpyxl uses 1-based indexing
start_row = 11
end_row = 96
start_col_idx = 14 # Column N is the 14th column
end_col_idx = 71   # Column BS is the 71st column (BS = 2*26 + 19 = 52 + 19 = 71)

for row in range(start_row, end_row + 1):
    # Extract row metadata (Columns A-I, which are indices 1-9)
    # A=1, B=2, C=3, D=4, E=5, F=6, G=7, H=8, I=9
    # Values for: '기관코드', '기관명', '기관구분', '소속기관차수', '중분류', '소분류', '기구유형', '보임직급', '상호이체보임직급'
    row_metadata = [ws.cell(row=row, column=c).value for c in range(1, 10)]
    
    for col in range(start_col_idx, end_col_idx + 1):
        # Extract column metadata (Rows 6, 7, 8 for the current column)
        # Values for: '직군', '계급_및_직급', '직렬'
        col_metadata = [
            ws.cell(row=6, column=col).value,
            ws.cell(row=7, column=col).value,
            ws.cell(row=8, column=col).value
        ]
        
        # Get the cell value (기준정원 count)
        cell_value = ws.cell(row=row, column=col).value
        
        # Process only if cell_value is a valid number > 0
        if isinstance(cell_value, (int, float)) and cell_value > 0:
            count = int(cell_value)
            
            # Create the row data
            # Combine row_metadata + col_metadata + [1] (as count is now distributed)
            new_row = row_metadata + col_metadata + [1]
            
            # Append this row 'count' times
            for _ in range(count):
                data.append(new_row)

# Create DataFrame
df = pd.DataFrame(data, columns=columns)

# Save to Excel
try:
    df.to_excel(output_file, index=False)
    print(f"Successfully created {output_file} with {len(df)} rows.")
except Exception as e:
    print(f"Error saving output file: {e}")
